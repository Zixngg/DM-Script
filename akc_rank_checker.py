import undetected_chromedriver as uc
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException, WebDriverException
import pandas as pd
from urllib.parse import urlparse, urlencode, parse_qs, urljoin
from datetime import datetime
import time
import random
import os
import re

# ========== CONFIGURATION ==========
INPUT_EXCEL = 'Book1.xlsx'  # first column contains search terms
OUTPUT_SHEET_NAME = 'AKC Rankings'  # results sheet inside INPUT_EXCEL
PAGE_READY_TIMEOUT_SECONDS = 10
GOOGLE_RESULTS_PAGES = 1  # 1 page = top 10, 2 pages = top 20, etc.
RESULTS_PER_PAGE = 10
VERBOSE = False  # Toggle detailed [DEBUG] logs

# Domain(s) to detect as the AKC site
TARGET_DOMAINS = ['sg-akc.com']


def setup_driver():
    # Use undetected-chromedriver by default to avoid Google detection
    try:
        options = uc.ChromeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        # Add user agent to appear more human-like
        options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        uc.Chrome.__del__ = lambda self: None
        driver = uc.Chrome(options=options)
        driver.maximize_window()
        if VERBOSE: print("[DEBUG] Using undetected-chromedriver")
        return driver
    except Exception as e:
        if VERBOSE: print(f"[DEBUG] Failed to use undetected-chromedriver: {e}, trying regular Selenium")
        # Fallback to regular Selenium
        options = webdriver.ChromeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        driver = webdriver.Chrome(options=options)
        driver.maximize_window()
        return driver


def wait_until_ready(driver, timeout_seconds=PAGE_READY_TIMEOUT_SECONDS):
    try:
        WebDriverWait(driver, timeout_seconds).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
    except TimeoutException:
        pass
    time.sleep(random.uniform(0.5, 1.2))


def is_unusual_traffic(driver) -> bool:
    try:
        return 'unusual traffic' in driver.page_source.lower()
    except Exception:
        return False


def get_base_domain(url: str) -> str:
    try:
        netloc = urlparse(url).netloc.lower()
        result = netloc[4:] if netloc.startswith('www.') else netloc
        if url and not result:
            if VERBOSE: print(f"[DEBUG] get_base_domain: url='{url[:100]}', netloc='{netloc}', result='{result}'")
        return result
    except Exception as e:
        if VERBOSE: print(f"[DEBUG] get_base_domain error: url='{url[:100] if url else None}', error={str(e)[:50]}")
        return ''


def is_target_domain(url: str) -> bool:
    base = get_base_domain(url)
    if not base:
        return False
    # Check if domain matches or contains target domain
    for td in TARGET_DOMAINS:
        match_exact = (base == td)
        match_endswith = base.endswith('.' + td)
        match_contains = (td in base)
        if match_exact or match_endswith or match_contains:
            print(f"[INFO] Domain match: base='{base}', target='{td}', exact={match_exact}, endswith={match_endswith}, contains={match_contains}")
            return True
    return False


def _normalize_google_result_href(href: str) -> str:
    # Convert Google redirect links like /url?q=https://target... to the target URL
    try:
        if not href:
            return href
        # Handle relative URLs
        if href.startswith('/url?'):
            absolute = urljoin('https://www.google.com', href)
            parsed = urlparse(absolute)
            q = parse_qs(parsed.query).get('q', [])
            if q:
                return q[0]
        # Handle absolute Google URLs with /url?q=
        elif 'google.' in href and '/url?' in href:
            parsed = urlparse(href)
            q = parse_qs(parsed.query).get('q', [])
            if q:
                return q[0]
        return href
    except Exception:
        return href


def google_search_collect_results(driver, query: str, pages: int) -> list[str]:
    if VERBOSE: print(f"[DEBUG] google_search_collect_results called: query='{query}', pages={pages}")
    collected_urls: list[str] = []
    
    try:
        for page_index in range(pages):
            if VERBOSE: print(f"[DEBUG] Processing page {page_index + 1} of {pages}")
            start = page_index * RESULTS_PER_PAGE
            params = {'q': query, 'start': start, 'num': RESULTS_PER_PAGE, 'hl': 'en'}
            google_url = f"https://www.google.com/search?{urlencode(params)}"
            if VERBOSE: print(f"[DEBUG] Loading Google URL: {google_url}")

            try:
                driver.set_page_load_timeout(20)
                driver.get(google_url)
                if VERBOSE: print(f"[DEBUG] Page loaded successfully")
            except TimeoutException as e:
                if VERBOSE: print(f"[DEBUG] Page load timeout: {e}")
                try:
                    driver.execute_script('window.stop()')
                except Exception:
                    pass
            except WebDriverException as e:
                if VERBOSE: print(f"[DEBUG] WebDriverException: {e}")
                # transient; try next page or stop
                break

            wait_until_ready(driver)
            if VERBOSE: print(f"[DEBUG] Page ready, waiting for results...")
            
            # Add a small random delay to appear more human-like
            time.sleep(random.uniform(1.5, 3.0))

            if is_unusual_traffic(driver):
                # Back off instead of hammering
                wait_time = random.randint(60, 120)
                if VERBOSE: print(f"[DEBUG] Unusual traffic detected, waiting {wait_time} seconds...")
                if VERBOSE: print(f"[DEBUG] This may be a CAPTCHA. Please check the browser window.")
                time.sleep(wait_time)
                
                # Try to refresh the page after waiting
                try:
                    driver.refresh()
                    wait_until_ready(driver)
                    time.sleep(3)
                    
                    # Check again if still blocked
                    if is_unusual_traffic(driver):
                        if VERBOSE: print(f"[DEBUG] Still blocked after refresh. You may need to solve CAPTCHA manually.")
                        if VERBOSE: print(f"[DEBUG] Waiting additional 30 seconds for manual intervention...")
                        time.sleep(30)
                        driver.refresh()
                        wait_until_ready(driver)
                        time.sleep(3)
                except Exception as e:
                    if VERBOSE: print(f"[DEBUG] Error refreshing page: {e}")
                
                # Check one more time
                if is_unusual_traffic(driver):
                    if VERBOSE: print(f"[DEBUG] Page still shows unusual traffic. Skipping this page.")
                    continue

            # Wait for results to load - try waiting for specific elements
            try:
                WebDriverWait(driver, 5).until(
                    lambda d: len(d.find_elements(By.CSS_SELECTOR, 'div.g, div.yuRUbf, cite')) > 0
                )
                if VERBOSE: print(f"[DEBUG] Results loaded successfully")
            except TimeoutException:
                if VERBOSE: print(f"[DEBUG] Timeout waiting for results to load")
            
            # Wait a bit more for results to render
            time.sleep(2)

            # Debug: Check page structure
            if VERBOSE: print(f"\n[DEBUG] Page title: {driver.title}")
            if VERBOSE: print(f"[DEBUG] Page URL: {driver.current_url}")
            
            # Try to find ANY cite elements on the page
            all_cites = driver.find_elements(By.CSS_SELECTOR, 'cite')
            if VERBOSE: print(f"[DEBUG] Found {len(all_cites)} cite elements on page")
            if all_cites:
                for i, cite in enumerate(all_cites[:5], 1):  # Show first 5
                    try:
                        cite_text = cite.text.strip()
                        print(f"  Cite {i}: {cite_text[:100]}")
                    except:
                        pass
            
            # Try to find ANY links in search results
            all_links = driver.find_elements(By.CSS_SELECTOR, '#search a, #rso a')
            if VERBOSE: print(f"[DEBUG] Found {len(all_links)} links in search area")
            
            # Try multiple selectors to find result blocks
            result_blocks = []
            selectors_to_try = [
                'div.yuRUbf',  # Standard organic results
                'div[class*="yuRUbf"]',  # With class variations
                'div.g',  # Generic result container
                'div[data-sokoban-container]',  # Another container type
                'div[class*="g"]',  # Any div with g class
                'div#search div[class*="g"]',  # Results in search div
            ]
            
            for selector in selectors_to_try:
                blocks = driver.find_elements(By.CSS_SELECTOR, selector)
                if blocks:
                    result_blocks = blocks
                    if VERBOSE: print(f"[DEBUG] Found {len(blocks)} result blocks using selector: {selector}")
                    break
            
            if not result_blocks:
                if VERBOSE: print(f"[DEBUG] No result blocks found with any selector")
                # Try to find div.byrV5b which contains the cite
                byrV5b_blocks = driver.find_elements(By.CSS_SELECTOR, 'div.byrV5b')
                if VERBOSE: print(f"[DEBUG] Found {len(byrV5b_blocks)} div.byrV5b blocks")
                if byrV5b_blocks:
                    result_blocks = byrV5b_blocks
                    if VERBOSE: print(f"[DEBUG] Using div.byrV5b blocks as result containers")
            
            page_urls = []
            seen = set()
            
            if VERBOSE: print(f"[DEBUG] Processing {len(result_blocks)} result blocks...")
            
            for block_idx, block in enumerate(result_blocks, 1):
                try:
                    if not block.is_displayed():
                        print(f"[DEBUG] Block {block_idx}: Not displayed, skipping")
                        continue
                
                    print(f"[DEBUG] ===== Block {block_idx} ======")
                    
                    # Extract from cite element first (most reliable for domain matching)
                    # The cite element contains the URL like "https://www.sg-akc.com › category › food-safety-courses"
                    href = None
                    cite = None
                    try:
                        # Try multiple cite selectors
                        cite_selectors = [
                            'cite', 
                            'cite.qLRx3b', 
                            'cite[class*="qLRx3b"]', 
                            'cite[class*="tjvcx"]',
                            'cite[class*="GvPZzd"]',
                            'div.byrV5b cite',
                            '.byrV5b cite'
                        ]
                        for cite_sel in cite_selectors:
                            try:
                                cite = block.find_element(By.CSS_SELECTOR, cite_sel)
                                if cite:
                                    cite_text_check = cite.text.strip()
                                    print(f"[DEBUG] Block {block_idx}: Found cite with selector '{cite_sel}', text: '{cite_text_check[:100]}'")
                                    if cite_text_check:
                                        break
                            except Exception as e:
                                print(f"[DEBUG] Block {block_idx}: Selector '{cite_sel}' failed: {str(e)[:50]}")
                                continue
                        
                        if not cite:
                            print(f"[DEBUG] Block {block_idx}: No cite element found with any selector")
                        
                        if cite:
                            # Extract URL from cite element
                            # The cite structure: <cite>https://www.sg-akc.com<span> › category › food-safety-courses</span></cite>
                            
                            print(f"[DEBUG] Block {block_idx}: Starting cite extraction...")
                            
                            # Method 1: Try textContent which includes all text nodes
                            cite_text = None
                            try:
                                cite_text = cite.get_attribute('textContent') or cite.get_attribute('innerText')
                                if cite_text:
                                    cite_text = cite_text.strip()
                                    print(f"[DEBUG] Block {block_idx}: Method 1 (textContent/innerText): '{cite_text[:150]}'")
                            except Exception as e:
                                print(f"[DEBUG] Block {block_idx}: Method 1 failed: {str(e)[:50]}")
                            
                            # Method 2: Fallback to .text property
                            if not cite_text:
                                try:
                                    cite_text = cite.text.strip()
                                    print(f"[DEBUG] Block {block_idx}: Method 2 (.text property): '{cite_text[:150]}'")
                                except Exception as e:
                                    print(f"[DEBUG] Block {block_idx}: Method 2 failed: {str(e)[:50]}")
                            
                            # Method 3: Extract directly from outerHTML using regex
                            if not cite_text or not cite_text.startswith('http'):
                                try:
                                    outer_html = cite.get_attribute('outerHTML') or ''
                                    print(f"[DEBUG] Block {block_idx}: Method 3 (outerHTML): '{outer_html[:200]}'")
                                    # Extract URL pattern: https://www.sg-akc.com
                                    url_pattern = r'https?://[^\s<>&"]+'
                                    url_match = re.search(url_pattern, outer_html)
                                    if url_match:
                                        cite_text = url_match.group(0)
                                        print(f"[DEBUG] Block {block_idx}: Method 3 extracted: '{cite_text}'")
                                except Exception as e:
                                    print(f"[DEBUG] Block {block_idx}: Method 3 failed: {str(e)[:50]}")
                            
                            if cite_text:
                                print(f"[DEBUG] Block {block_idx}: Processing cite_text: '{cite_text[:150]}'")
                                
                                # Extract the base URL (everything before the first ›, space, or <)
                                # Format: "https://www.sg-akc.com › category › food-safety-courses"
                                # Or: "https://www.sg-akc.com" (without path)
                                
                                # Remove any HTML entities
                                cite_text = cite_text.replace('&nbsp;', ' ').replace('&amp;', '&').strip()
                                
                                # Extract base URL - everything before first ›, <, or space (if not part of URL)
                                if '›' in cite_text:
                                    base_url = cite_text.split('›')[0].strip()
                                    print(f"[DEBUG] Block {block_idx}: Split by ›, base_url: '{base_url}'")
                                elif '<' in cite_text:
                                    base_url = cite_text.split('<')[0].strip()
                                    print(f"[DEBUG] Block {block_idx}: Split by <, base_url: '{base_url}'")
                                elif ' ' in cite_text:
                                    # If it starts with http, take first word (the URL)
                                    if cite_text.startswith('http'):
                                        base_url = cite_text.split()[0]
                                        print(f"[DEBUG] Block {block_idx}: Split by space (http detected), base_url: '{base_url}'")
                                    else:
                                        base_url = cite_text
                                        print(f"[DEBUG] Block {block_idx}: Split by space (no http), base_url: '{base_url}'")
                                else:
                                    base_url = cite_text.strip()
                                    print(f"[DEBUG] Block {block_idx}: No split needed, base_url: '{base_url}'")
                                
                                # Normalize to full URL - just need domain for ranking
                                if base_url.startswith('http'):
                                    href = base_url
                                    print(f"[DEBUG] Block {block_idx}: href normalized (http): '{href[:100]}'")
                                elif '.' in base_url:
                                    href = f"https://{base_url}"
                                    print(f"[DEBUG] Block {block_idx}: href normalized (added https): '{href[:100]}'")
                                else:
                                    href = None
                                    print(f"[DEBUG] Block {block_idx}: Could not normalize href from: '{base_url}'")
                                
                                if href:
                                    domain = get_base_domain(href)
                                    is_akc = is_target_domain(href)
                                    print(f"[DEBUG] Block {block_idx}: Final href: '{href[:100]}', domain: '{domain}', is_akc: {is_akc}")
                            else:
                                print(f"[DEBUG] Block {block_idx}: No cite_text extracted")
                    except Exception as e:
                        print(f"[DEBUG] Error extracting cite: {e}")
                        pass
                    
                    # Fallback: try to get href from anchor if cite didn't work
                    if not href or not href.startswith('http'):
                        print(f"[DEBUG] Block {block_idx}: Trying anchor fallback...")
                        try:
                            # Try multiple anchor selectors
                            a = None
                            anchor_selectors = ['a', 'a[href]', 'h3 a']
                            for anchor_sel in anchor_selectors:
                                try:
                                    a = block.find_element(By.CSS_SELECTOR, anchor_sel)
                                    if a:
                                        href_attr_raw = a.get_attribute('href') or ''
                                        print(f"[DEBUG] Block {block_idx}: Found anchor with selector '{anchor_sel}', raw href: '{href_attr_raw[:100]}'")
                                        break
                                except Exception as e:
                                    print(f"[DEBUG] Block {block_idx}: Anchor selector '{anchor_sel}' failed: {str(e)[:50]}")
                                    continue
                            
                            if a:
                                href_attr = a.get_attribute('href') or ''
                                if href_attr:
                                    href = _normalize_google_result_href(href_attr)
                                    print(f"[DEBUG] Block {block_idx}: Normalized anchor href: '{href[:100]}'")
                            else:
                                print(f"[DEBUG] Block {block_idx}: No anchor found")
                        except Exception as e:
                            print(f"[DEBUG] Block {block_idx}: Error extracting anchor: {e}")
                            pass
                    
                    if not href or not href.startswith('http'):
                        if href:
                            if VERBOSE: print(f"[DEBUG] Block {block_idx}: Skipping invalid href: '{href[:80]}'")
                        else:
                            if VERBOSE: print(f"[DEBUG] Block {block_idx}: No href extracted")
                        continue
                    
                    domain = get_base_domain(href)
                    if VERBOSE: print(f"[DEBUG] Block {block_idx}: Extracted domain: '{domain}'")
                    
                    # Check if it's AKC
                    is_akc = is_target_domain(href)
                    if VERBOSE: print(f"[DEBUG] Block {block_idx}: Is AKC domain? {is_akc}")
                    
                    # Filter out Google-owned or non-organic domains
                    if any(x in domain for x in ['google.', 'gstatic.', 'youtube.', 'webcache.googleusercontent']):
                        if VERBOSE: print(f"[DEBUG] Block {block_idx}: Skipping Google domain: {domain}")
                        continue
                    if '/maps' in href or '/search?' in href:
                        if VERBOSE: print(f"[DEBUG] Block {block_idx}: Skipping maps/search link: {href[:80]}")
                        continue
                    
                    # Deduplicate by domain for organic results
                    if domain and domain not in seen:
                        seen.add(domain)
                        page_urls.append(href)
                        if VERBOSE: print(f"[DEBUG] Block {block_idx}: ✓ Added URL #{len(page_urls)}: domain='{domain}', href='{href[:100]}', is_akc={is_akc}")
                    else:
                        if VERBOSE: print(f"[DEBUG] Block {block_idx}: ✗ Skipping duplicate domain: '{domain}'")
                except Exception:
                    continue

            collected_urls.extend(page_urls)
            if VERBOSE: print(f"[DEBUG] Page {page_index + 1}: Collected {len(page_urls)} URLs (total: {len(collected_urls)})")
            
            # Save screenshot for debugging if no URLs found
            if not page_urls and page_index == 0:
                try:
                    screenshot_path = f"debug_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                    driver.save_screenshot(screenshot_path)
                    if VERBOSE: print(f"[DEBUG] Saved screenshot to {screenshot_path}")
                except Exception as e:
                    if VERBOSE: print(f"[DEBUG] Could not save screenshot: {e}")
            
            time.sleep(random.uniform(0.8, 1.6))
    
    except Exception as e:
        if VERBOSE: print(f"[DEBUG] ERROR in google_search_collect_results: {e}")
        import traceback
        if VERBOSE: print(f"[DEBUG] Traceback: {traceback.format_exc()}")
    
    if VERBOSE: print(f"[DEBUG] google_search_collect_results returning {len(collected_urls)} URLs")
    return collected_urls


def find_rank_for_query(driver, query: str, pages: int) -> int | None:
    urls = google_search_collect_results(driver, query, pages)
    if VERBOSE: print(f"\n[DEBUG] Query: '{query}'")
    if VERBOSE: print(f"[DEBUG] Found {len(urls)} URLs:")
    for idx, url in enumerate(urls[:10], 1):  # Show first 10
        domain = get_base_domain(url)
        is_match = is_target_domain(url)
        if VERBOSE: print(f"  {idx}. {domain} -> {url[:80]}... {'✓ MATCH' if is_match else ''}")
    
    for index, url in enumerate(urls, start=1):
        if is_target_domain(url):
            print(f"[INFO] Found AKC at rank {index}")
            return index
    if VERBOSE: print(f"[DEBUG] AKC not found in {len(urls)} results")
    return None


def read_search_terms_from_excel(path: str) -> list[str]:
    df = pd.read_excel(path)
    if df.empty:
        return []
    return df.iloc[:, 0].dropna().astype(str).tolist()


def write_results_to_excel(path: str, rows: list[dict]):
    # Append-or-create sheet `AKC Rankings`
    if not rows:
        return

    # Try to read existing file, if permission error, create new file
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        # Try to load existing workbook
        try:
            wb = load_workbook(path)
        except PermissionError:
            print(f"[WARN] Cannot open {path} - file may be open in Excel.")
            print(f"[WARN] Creating backup file: {path.replace('.xlsx', '_rankings.xlsx')}")
            # Create new file with just rankings
            output_path = path.replace('.xlsx', '_rankings.xlsx')
            new_df = pd.DataFrame(rows, columns=['Search term', 'Results Ranking', 'Date'])
            new_df.to_excel(output_path, sheet_name=OUTPUT_SHEET_NAME, index=False, engine='openpyxl')
            print(f"[INFO] Results saved to {output_path}")
            return
        
        # Read existing sheet if it exists
        try:
            existing = pd.read_excel(path, sheet_name=OUTPUT_SHEET_NAME)
        except Exception:
            existing = pd.DataFrame()

        # Combine with new data
        new_df = pd.DataFrame(rows, columns=['Search term', 'Results Ranking', 'Date'])
        combined = pd.concat([existing, new_df], ignore_index=True)
        
        # Remove old sheet and write new one
        if OUTPUT_SHEET_NAME in wb.sheetnames:
            wb.remove(wb[OUTPUT_SHEET_NAME])
        
        # Create new sheet
        ws = wb.create_sheet(OUTPUT_SHEET_NAME)
        
        # Write headers
        ws.append(['Search term', 'Results Ranking', 'Date'])
        
        # Write data
        for _, row in combined.iterrows():
            ws.append([row['Search term'], row['Results Ranking'], row['Date']])
        
        wb.save(path)
        wb.close()
        print(f"[INFO] Results saved to sheet '{OUTPUT_SHEET_NAME}' in {path}")
        
    except Exception as e:
        print(f"[ERROR] Failed to write to Excel: {e}")
        # Fallback: create CSV
        csv_path = path.replace('.xlsx', '_rankings.csv')
        new_df = pd.DataFrame(rows, columns=['Search term', 'Results Ranking', 'Date'])
        new_df.to_csv(csv_path, index=False)
        print(f"[INFO] Results saved to CSV: {csv_path}")


def main():
    terms = read_search_terms_from_excel(INPUT_EXCEL)
    if not terms:
        print('No search terms found in the first column of the Excel file.')
        return

    driver = setup_driver()
    results_rows: list[dict] = []
    today = datetime.now().strftime('%Y-%m-%d')

    try:
        for term in terms:
            rank = find_rank_for_query(driver, term, GOOGLE_RESULTS_PAGES)
            results_rows.append({
                'Search term': term,
                'Results Ranking': rank if rank is not None else 'Not Found',
                'Date': today,
            })
            # small human-like pause
            time.sleep(random.uniform(1.0, 2.0))
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    write_results_to_excel(INPUT_EXCEL, results_rows)
    print(f"Wrote {len(results_rows)} rows to sheet '{OUTPUT_SHEET_NAME}' in {INPUT_EXCEL}")


if __name__ == '__main__':
    main()


